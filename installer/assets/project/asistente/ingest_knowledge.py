"""
ingest_knowledge.py — Inyector de conocimiento desde disco
===========================================================
Lee archivos/carpetas del disco duro e inyecta el contenido
como patrones de conocimiento en la KB.

USO:
  python ingest_knowledge.py <path>
  python ingest_knowledge.py <path> --domain <nombre>
  python ingest_knowledge.py <path> --domain <nombre> --type fact
  python ingest_knowledge.py <path> --domain <nombre> --tags tag1,tag2
  python ingest_knowledge.py <path> --preview    (solo muestra que haria, no guarda)

EJEMPLOS:
  python ingest_knowledge.py C:/datos/contabilidad/
  python ingest_knowledge.py C:/docs/manual_finanzas.txt --domain finanzas
  python ingest_knowledge.py C:/kb/ --domain general --tags manual,empresa

FORMATOS SOPORTADOS:
  .txt, .md, .py, .json, .csv, .log
  .docx  (requiere: pip install python-docx)
  .pdf   (requiere: pip install pdfplumber)
  .xlsx  (requiere: pip install openpyxl)
"""

import argparse
import json
import sys
import re
from pathlib import Path
from datetime import datetime

# Agregar el proyecto al path
PROJECT_DIR = Path(__file__).parent
sys.path.insert(0, str(PROJECT_DIR))

SUPPORTED_EXTENSIONS = {".txt", ".md", ".py", ".json", ".csv", ".log", ".docx", ".pdf", ".xlsx"}
CHUNK_SIZE = 800       # chars por chunk
CHUNK_OVERLAP = 100    # overlap entre chunks para no perder contexto


# ── Lectura de archivos ────────────────────────────────────────

def read_txt(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8", errors="replace")
    except Exception as e:
        return ""


def read_json(path: Path) -> str:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            return json.dumps(data, indent=2, ensure_ascii=False)[:5000]
        elif isinstance(data, list):
            return "\n".join(str(item) for item in data[:50])
        return str(data)[:2000]
    except Exception:
        return read_txt(path)


def read_docx(path: Path) -> str:
    try:
        from docx import Document
        doc = Document(str(path))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except ImportError:
        print(f"  ⚠  python-docx no instalado. Instalar: pip install python-docx")
        return ""
    except Exception as e:
        print(f"  ⚠  Error leyendo {path.name}: {e}")
        return ""


def read_pdf(path: Path) -> str:
    try:
        import pdfplumber
        with pdfplumber.open(str(path)) as pdf:
            return "\n".join(
                page.extract_text() or ""
                for page in pdf.pages[:30]  # máximo 30 páginas
            )
    except ImportError:
        print(f"  ⚠  pdfplumber no instalado. Instalar: pip install pdfplumber")
        return ""
    except Exception as e:
        print(f"  ⚠  Error leyendo {path.name}: {e}")
        return ""


def read_xlsx(path: Path) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        lines = []
        for sheet in wb.sheetnames[:3]:
            ws = wb[sheet]
            lines.append(f"[Hoja: {sheet}]")
            for row in ws.iter_rows(max_row=100, values_only=True):
                row_text = " | ".join(str(c) for c in row if c is not None)
                if row_text.strip():
                    lines.append(row_text)
        return "\n".join(lines)
    except ImportError:
        print(f"  ⚠  openpyxl no instalado. Instalar: pip install openpyxl")
        return ""
    except Exception as e:
        print(f"  ⚠  Error leyendo {path.name}: {e}")
        return ""


def read_file(path: Path) -> str:
    """Despacha al lector correcto según extensión."""
    ext = path.suffix.lower()
    if ext == ".json":
        return read_json(path)
    elif ext == ".docx":
        return read_docx(path)
    elif ext == ".pdf":
        return read_pdf(path)
    elif ext == ".xlsx":
        return read_xlsx(path)
    else:
        return read_txt(path)


# ── Chunking ──────────────────────────────────────────────────

def chunk_text(text: str, chunk_size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> list:
    """Divide el texto en chunks con overlap para no perder contexto en los bordes."""
    if not text or not text.strip():
        return []
    text = re.sub(r'\n{3,}', '\n\n', text.strip())
    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end]
        if chunk.strip():
            chunks.append(chunk.strip())
        start = end - overlap
    return chunks


# ── Detección de dominio ───────────────────────────────────────

def detect_domain_for_content(content: str, filename: str) -> str:
    """Detecta dominio usando domain_detector. Fallback: 'general'."""
    try:
        from domain_detector import detect
        # Combinar nombre del archivo + primeros 500 chars del contenido
        search_text = f"{filename} {content[:500]}"
        domain = detect(search_text)
        return domain
    except Exception:
        return "general"


# ── Registro en KB ────────────────────────────────────────────

def ingest_chunk(chunk: str, domain: str, source_file: str,
                 chunk_idx: int, entry_type: str, tags: list,
                 preview: bool = False) -> bool:
    """Registra un chunk de texto como entrada en la KB."""
    context_key = f"ingested_{Path(source_file).stem}_{chunk_idx}"
    solution = {
        "strategy":    "ingested_content",
        "content":     chunk,
        "source_file": source_file,
        "chunk_index": chunk_idx,
        "notes":       f"Ingestado desde: {Path(source_file).name} (chunk {chunk_idx})",
        "ingested_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    all_tags = list(set(["ingested", domain] + tags))

    if preview:
        print(f"    [{chunk_idx}] dominio={domain} | {chunk[:80].replace(chr(10),' ')}...")
        return True

    try:
        from knowledge_base import register_pattern
        register_pattern(
            domain=domain,
            context_key=context_key,
            solution=solution,
            tags=all_tags,
            entry_type=entry_type,
        )
        return True
    except Exception as e:
        print(f"    ✗ Error registrando chunk {chunk_idx}: {e}")
        return False


# ── Procesamiento de archivos ──────────────────────────────────

def process_file(path: Path, domain: str | None, entry_type: str,
                 tags: list, preview: bool) -> tuple:
    """Procesa un archivo. Retorna (chunks_total, chunks_ok)."""
    print(f"\n  📄 {path.name}")
    content = read_file(path)
    if not content or not content.strip():
        print(f"     ⚠  vacío o no se pudo leer")
        return 0, 0

    # Auto-detectar dominio si no se especificó
    actual_domain = domain or detect_domain_for_content(content, path.name)
    if not domain:
        print(f"     🔍 dominio auto-detectado: {actual_domain}")

    chunks = chunk_text(content)
    print(f"     chunks: {len(chunks)} | dominio: {actual_domain} | tipo: {entry_type}")

    ok = 0
    for i, chunk in enumerate(chunks):
        if ingest_chunk(chunk, actual_domain, str(path), i + 1, entry_type, tags, preview):
            ok += 1

    # Aprender las keywords de este contenido para el dominio
    if not preview and actual_domain != "general":
        try:
            from domain_detector import auto_learn_from_session
            auto_learn_from_session(actual_domain, content[:2000])
        except Exception:
            pass

    return len(chunks), ok


def collect_files(path: Path) -> list:
    """Recolecta todos los archivos soportados en path (file o directorio)."""
    if path.is_file():
        if path.suffix.lower() in SUPPORTED_EXTENSIONS:
            return [path]
        else:
            print(f"⚠  Formato no soportado: {path.suffix}")
            return []
    elif path.is_dir():
        files = []
        for f in sorted(path.rglob("*")):
            if f.is_file() and f.suffix.lower() in SUPPORTED_EXTENSIONS:
                # Ignorar archivos ocultos, __pycache__, .git
                if not any(part.startswith('.') or part == '__pycache__'
                           for part in f.parts):
                    files.append(f)
        return files
    else:
        print(f"✗ Path no encontrado: {path}")
        return []


# ── Main ──────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Inyecta conocimiento desde archivos/carpetas a la KB",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__
    )
    parser.add_argument("path",        help="Archivo o carpeta a ingestar")
    parser.add_argument("--domain",    help="Dominio destino (auto-detectado si no se especifica)")
    parser.add_argument("--type",      default="pattern", choices=["pattern", "fact"],
                        help="Tipo de entrada: pattern (default) o fact")
    parser.add_argument("--tags",      default="", help="Tags adicionales separados por coma")
    parser.add_argument("--preview",   action="store_true",
                        help="Mostrar qué haría sin guardar nada")
    args = parser.parse_args()

    path = Path(args.path)
    tags = [t.strip() for t in args.tags.split(",") if t.strip()] if args.tags else []

    print(f"\n{'='*60}")
    print(f"  Motor_Inteligente_IA — Ingestor de Conocimiento")
    print(f"{'='*60}")
    print(f"  Path:    {path}")
    print(f"  Dominio: {args.domain or '(auto-detectar)'}")
    print(f"  Tipo:    {args.type}")
    print(f"  Tags:    {tags or '(ninguno extra)'}")
    print(f"  Modo:    {'PREVIEW (no guarda)' if args.preview else 'REAL (guarda en KB)'}")
    print(f"{'='*60}")

    files = collect_files(path)
    if not files:
        print("\n✗ No se encontraron archivos soportados.")
        sys.exit(1)

    print(f"\nArchivos encontrados: {len(files)}")
    total_chunks = 0
    total_ok     = 0

    for f in files:
        c, ok = process_file(f, args.domain, args.type, tags, args.preview)
        total_chunks += c
        total_ok     += ok

    print(f"\n{'='*60}")
    if args.preview:
        print(f"  PREVIEW: {total_chunks} chunks habrían sido ingestados desde {len(files)} archivo(s)")
        print(f"  Ejecutar sin --preview para guardar.")
    else:
        print(f"  ✓ {total_ok}/{total_chunks} chunks ingestados desde {len(files)} archivo(s)")
        if total_ok > 0:
            print(f"  KB actualizada. Verificar con: python knowledge_base.py list-domains")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
